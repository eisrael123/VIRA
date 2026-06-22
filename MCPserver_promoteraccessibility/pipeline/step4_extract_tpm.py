"""Step 4: classify DESeq2 genes and attach CAGE seed-peak coordinates."""

from __future__ import annotations

import os
import re
from dataclasses import dataclass

from openpyxl import load_workbook

CNTL_COLUMN_PATTERN = r"_cntl\d+_"
TEST_COLUMN_PATTERN = r"_test\d+_"


@dataclass
class Step4Result:
    induced_plot: str
    suppressed_plot: str
    nonsig_plot: str
    induced_genes: str
    suppressed_genes: str
    nonsig_genes: str
    tsv_path: str


def _looks_like_number(value: str) -> bool:
    if value is None or value == "":
        return False
    try:
        float(value)
        return True
    except (TypeError, ValueError):
        return False


def _cell_to_str(value) -> str:
    if value is None:
        return ""
    text = str(value)
    return "" if text == "NA" else text


def _xlsx_to_tsv(xlsx_path: str, tsv_path: str) -> None:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    with open(tsv_path, "w") as out:
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                out.write("\t".join(_cell_to_str(cell) for cell in row) + "\n")
    workbook.close()


def extract_tpm(
    deseq2_xlsx: str,
    main_peaks: str,
    output_dir: str,
    log2fc: float = 0.5,
    padj: float = 0.005,
    min_cntl_tpm: float = 3.0,
    sample_label: str | None = None,
    cntl_pattern: str = CNTL_COLUMN_PATTERN,
    test_pattern: str = TEST_COLUMN_PATTERN,
) -> Step4Result:
    os.makedirs(output_dir, exist_ok=True)
    if sample_label is None:
        sample_label = re.sub(
            r"_test_vs_cntl_deseq2_results_genes\.xlsx$", "", os.path.basename(deseq2_xlsx)
        )

    tsv_path = os.path.join(output_dir, f"{sample_label}.tsv")
    induced_genes = os.path.join(output_dir, f"Induced_genes_{sample_label}.txt")
    suppressed_genes = os.path.join(output_dir, f"Suppressed_genes_{sample_label}.txt")
    nonsig_genes = os.path.join(output_dir, f"Non-sig_genes_{sample_label}.txt")

    _xlsx_to_tsv(deseq2_xlsx, tsv_path)

    induced: dict[str, str] = {}
    suppressed: dict[str, str] = {}
    nonsig: dict[str, str] = {}

    cntl_re = re.compile(cntl_pattern, re.IGNORECASE)
    test_re = re.compile(test_pattern, re.IGNORECASE)

    with open(tsv_path) as fh:
        header = fh.readline().rstrip("\n").split("\t")
        log2fc_index = padj_index = -1
        cntl_indices: list[int] = []
        test_indices: list[int] = []

        for i, name in enumerate(header):
            if name == "log2FoldChange":
                log2fc_index = i
            if name == "padj":
                padj_index = i
            if cntl_re.search(name):
                cntl_indices.append(i)
            elif test_re.search(name):
                test_indices.append(i)

        if log2fc_index == -1 or padj_index == -1 or not cntl_indices or not test_indices:
            raise ValueError(
                "Missing required DESeq2 columns: log2FoldChange, padj, and control/test TPM columns."
            )

        for line in fh:
            cols = line.rstrip("\n").split("\t")
            test_values = [float(cols[i]) for i in test_indices if i < len(cols) and _looks_like_number(cols[i])]
            if not test_values:
                continue
            test_avg = sum(test_values) / len(test_values)

            cntl_values = [float(cols[i]) for i in cntl_indices if i < len(cols) and _looks_like_number(cols[i])]
            if not cntl_values:
                continue
            cntl_avg = sum(cntl_values) / len(cntl_values)
            if cntl_avg < min_cntl_tpm:
                continue

            logfc_value = cols[log2fc_index] if log2fc_index < len(cols) else ""
            padj_value = cols[padj_index] if padj_index < len(cols) else ""
            if not padj_value or re.fullmatch(r"NA", padj_value, re.IGNORECASE):
                continue
            if not (_looks_like_number(logfc_value) and _looks_like_number(padj_value)):
                raise ValueError(f"logFC or padj is not numeric in line: {line!r}")

            gene = cols[0]
            tpm_str = repr(test_avg)
            logfc = float(logfc_value)
            padj_num = float(padj_value)
            if logfc < -log2fc and padj_num < padj:
                suppressed[gene] = tpm_str
            elif logfc > log2fc and padj_num < padj:
                induced[gene] = tpm_str
            else:
                nonsig[gene] = tpm_str

    _write_gene_list(induced_genes, induced)
    _write_gene_list(suppressed_genes, suppressed)
    _write_gene_list(nonsig_genes, nonsig)

    induced_plot = os.path.join(output_dir, f"{sample_label}_induced.txt")
    suppressed_plot = os.path.join(output_dir, f"{sample_label}_suppressed.txt")
    nonsig_plot = os.path.join(output_dir, f"{sample_label}_nonsig.txt")

    with open(main_peaks) as fh, open(induced_plot, "w") as out_ind, open(
        suppressed_plot, "w"
    ) as out_sup, open(nonsig_plot, "w") as out_non:
        for line in fh:
            cols = line.rstrip("\n").split("\t")
            peak_info = cols[3].split(";")
            seed_peak = f"{peak_info[1]}\t{peak_info[2]}"
            gene = cols[6]

            if gene in induced:
                out_ind.write("\t".join([cols[0], seed_peak, cols[5], gene, induced[gene]]) + "\n")
            elif gene in suppressed:
                out_sup.write("\t".join([cols[0], seed_peak, cols[5], gene, suppressed[gene]]) + "\n")
            elif gene in nonsig:
                out_non.write("\t".join([cols[0], seed_peak, cols[5], gene, nonsig[gene]]) + "\n")

    return Step4Result(
        induced_plot=induced_plot,
        suppressed_plot=suppressed_plot,
        nonsig_plot=nonsig_plot,
        induced_genes=induced_genes,
        suppressed_genes=suppressed_genes,
        nonsig_genes=nonsig_genes,
        tsv_path=tsv_path,
    )


def _write_gene_list(path: str, mapping: dict[str, str]) -> None:
    with open(path, "w") as out:
        for gene, tpm in mapping.items():
            out.write(f"{gene}\t{tpm}\n")
